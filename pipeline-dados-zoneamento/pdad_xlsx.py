# -*- coding: utf-8 -*-
"""
Leitor genérico das tabelas da PDAD-DF 2021 (planilha oficial da CODEPLAN,
Relatorio_DF_percentual-2021.xlsx / Relatorio_DF_total-2021.xlsx —
fontes-pesquisa-2026-08-27/). Cada aba da planilha é uma tabela ("A1",
"A51", "A79" etc.), com layout meio irregular:

  - às vezes 1 linha de cabeçalho: Local | Total | Categoria1 | Categoria2 | ...
  - às vezes 2 linhas: CategoriaGeral (mesclada) na linha de cima,
    Sim/Não (ou Total/Sim/Não) na linha de baixo, "Local" no canto
  - a 1ª linha de dado é sempre "DF" (total do Distrito Federal), seguida
    de uma linha por RA, na grafia própria da CODEPLAN (com abreviações
    tipo "Plano Piloto" em vez de "Brasília").

ler_coluna() acha a linha "Local" sozinho, monta o nome de cada coluna
(juntando a categoria de cima com a sub-coluna, se houver) e devolve só a
coluna pedida, já como dict {RA canônica: valor}.
"""
import openpyxl
from common import normaliza_ra, fonte

_wb_cache = {}


def _wb(nome_arquivo):
    if nome_arquivo not in _wb_cache:
        _wb_cache[nome_arquivo] = openpyxl.load_workbook(fonte(nome_arquivo), data_only=True)
    return _wb_cache[nome_arquivo]


def _linha_parece_subcabecalho(ws, r, max_col=12):
    """True se a linha `r` parece ser um sub-cabeçalho Sim/Não (1ª célula
    vazia, e as células seguintes são textos curtos, não number/lugar)."""
    if ws.cell(row=r, column=1).value not in (None, ""):
        return False
    vistos = 0
    for c in range(2, max_col + 1):
        v = ws.cell(row=r, column=c).value
        if v not in (None, ""):
            vistos += 1
            if not isinstance(v, str) or len(v) > 20:
                return False
    return vistos >= 1


def _monta_colunas(ws, linha_local, max_col=40):
    """Devolve (nomes_de_coluna, linha_onde_comecam_os_dados).

    Duas variações de layout aparecem na planilha:
      Tipo A: título / categoria (col1 vazia) / 'Local' + Sim,Não,Sim,Não...
              -> a linha 'Local' já traz os sub-rótulos; a categoria fica
                 uma linha ACIMA.
      Tipo B: título / 'Local' + categoria1,None,categoria2,None...
              / (col1 vazia) + Sim,Não,Sim,Não...
              -> a linha 'Local' traz as categorias; o sub-rótulo Sim/Não
                 vem uma linha ABAIXO.
    """
    if _linha_parece_subcabecalho(ws, linha_local + 1, max_col):
        cat = [ws.cell(row=linha_local, column=c).value for c in range(1, max_col + 1)]
        sub = [ws.cell(row=linha_local + 1, column=c).value for c in range(1, max_col + 1)]
        primeira_linha_dado = linha_local + 2
    else:
        sub = [ws.cell(row=linha_local, column=c).value for c in range(1, max_col + 1)]
        cat = [ws.cell(row=linha_local - 1, column=c).value for c in range(1, max_col + 1)] \
            if linha_local > 2 else [None] * max_col
        primeira_linha_dado = linha_local + 1

    last = None
    cat_ff = []
    for v in cat:
        if v not in (None, ""):
            last = v
        cat_ff.append(last)

    nomes = []
    for i, s in enumerate(sub):
        c = cat_ff[i]
        if s in (None, ""):
            nomes.append(c if (c and i > 0) else None)
        elif c and c != s and i > 0:
            nomes.append(f"{c}_{s}")
        else:
            nomes.append(str(s).strip())
    nomes[0] = "Local"
    return nomes, primeira_linha_dado


def _acha_linha_local(ws, max_row=6):
    for r in range(1, max_row + 1):
        if str(ws.cell(row=r, column=1).value).strip().lower() == "local":
            return r
    raise ValueError(f"não achei a linha 'Local' nas primeiras {max_row} linhas de {ws.title}")


def ler_tabela(sheet, arquivo="Relatorio_DF_percentual-2021.xlsx"):
    """Devolve (colunas, linhas) — colunas = lista de nomes já resolvidos,
    linhas = dict RA_bruta -> lista de valores alinhada às colunas
    (inclui a linha 'DF')."""
    ws = _wb(arquivo)[sheet]
    linha_local = _acha_linha_local(ws)
    colunas, primeira_linha_dado = _monta_colunas(ws, linha_local)
    linhas = {}
    r = primeira_linha_dado
    while True:
        nome = ws.cell(row=r, column=1).value
        if nome is None or str(nome).strip() == "":
            break
        vals = [ws.cell(row=r, column=c).value for c in range(1, len(colunas) + 1)]
        linhas[str(nome).strip()] = vals
        r += 1
    return colunas, linhas


def ler_coluna(sheet, nome_coluna, arquivo="Relatorio_DF_percentual-2021.xlsx"):
    """Devolve {RA canônica: valor} pra uma coluna específica de uma tabela
    da PDAD (ex.: ler_coluna('A75', 'Rede Geral (CAESB)_Sim'))."""
    colunas, linhas = ler_tabela(sheet, arquivo)
    if nome_coluna not in colunas:
        raise ValueError(f"coluna {nome_coluna!r} não existe em {sheet}. Colunas: {colunas}")
    idx = colunas.index(nome_coluna)
    out = {}
    for nome_bruto, vals in linhas.items():
        if nome_bruto.upper() == "DF":
            continue
        ra = normaliza_ra(nome_bruto)
        if ra is None:
            continue
        out[ra] = vals[idx]
    return out
